import openpyxl
from openpyxl import load_workbook
import re
import logging

# ─── Detailed import debug logger ───────────────────────────────────────────────
_log = logging.getLogger('import_debug')
if not _log.handlers:
    _h = logging.FileHandler('import_debug.log', encoding='utf-8')
    _h.setFormatter(logging.Formatter('%(asctime)s  %(levelname)s  %(message)s'))
    _log.addHandler(_h)
    _log.setLevel(logging.DEBUG)
# ────────────────────────────────────────────────────────────────────────────────


class ImportService:
    """Service for importing data from external Excel files"""
    
    def __init__(self, excel_service):
        self.excel_service = excel_service
    
    def import_from_file(self, file_path):
        """
        Import data from external Excel file (like hours-2.xlsx)
        Returns dict with success status and imported data summary
        """
        _log.info('======================================================')
        _log.info(f'import_from_file START  file={file_path!r}')
        _log.info(f'  excel_service file_path={self.excel_service.file_path!r}')
        try:
            wb = load_workbook(file_path)
            _log.info(f'  Source workbook sheets: {wb.sheetnames}')
            # map sheet names to worksheets case-insensitively
            sheets = {name.lower().strip(): name for name in wb.sheetnames}
            def _find_sheet_key(key):
                # find a sheet name that contains the key
                for k, orig in sheets.items():
                    if key in k:
                        return orig
                return None
            result = {
                'success': True,
                'teachers': {'added': 0, 'updated': 0},
                'groups': {'added': 0, 'updated': 0},
                'subjects': {'added': 0, 'updated': 0},
                'errors': []
            }

            # Import Configuration (if provided) -> merge into current config
            conf_name = _find_sheet_key('configuration')
            if conf_name:
                try:
                    conf_ws = wb[conf_name]
                    cfg_updates = {}
                    for row in conf_ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        key = str(row[0]).strip()
                        val = row[1] if len(row) > 1 else ''
                        cfg_updates[key] = val
                    if cfg_updates:
                        # merge with existing config and save
                        try:
                            existing = self.excel_service.get_config() or {}
                            existing.update(cfg_updates)
                            self.excel_service.save_config(existing)
                            result['config'] = cfg_updates
                        except Exception:
                            result['errors'].append('Failed to save imported configuration')
                except Exception:
                    result['errors'].append('Failed to read Configuration sheet')

            # Import Time Slots (if provided)
            ts_name = _find_sheet_key('time slot')
            if ts_name:
                try:
                    ts_ws = wb[ts_name]
                    time_slots = {}
                    for row in ts_ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        try:
                            lesson_num = int(row[0])
                        except Exception:
                            # skip invalid lesson numbers
                            continue
                        time_range = row[1] if len(row) > 1 and row[1] is not None else ''
                        if time_range:
                            time_slots[lesson_num] = str(time_range)

                    if time_slots:
                        try:
                            self.excel_service.save_time_slots(time_slots)
                            result['time_slots'] = {'imported': len(time_slots)}
                        except Exception:
                            result['errors'].append('Failed to save imported Time Slots')
                except Exception:
                    result['errors'].append('Failed to read Time Slots sheet')
            
            # Import Teachers (case-insensitive match)
            tname = _find_sheet_key('teacher')
            _log.info(f'  Teachers sheet found as: {tname!r}')
            if tname:
                teachers_before = self.excel_service.get_teachers()
                _log.info(f'  Teachers BEFORE _import_teachers: {len(teachers_before)} — {[t["name"] for t in teachers_before]}')
                teachers_stats = self._import_teachers(wb[tname])
                result['teachers'] = teachers_stats
                teachers_after = self.excel_service.get_teachers()
                _log.info(f'  Teachers AFTER _import_teachers: {len(teachers_after)} — {[t["name"] for t in teachers_after]}')
            else:
                result['errors'].append('Teachers sheet not found')
            
            # Import Rooms (case-insensitive, standalone Rooms sheet)
            rname = _find_sheet_key('room')
            _log.info(f'  Rooms sheet found as: {rname!r}')
            if rname:
                try:
                    rooms_ws = wb[rname]
                    rooms_list = []
                    for row in rooms_ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        rooms_list.append({
                            'name': str(row[0]).strip(),
                            'description': str(row[1]).strip() if len(row) > 1 and row[1] else ''
                        })
                    _log.info(f'  Rooms parsed: {len(rooms_list)} — {[r["name"] for r in rooms_list]}')
                    self.excel_service.save_rooms(rooms_list)
                    result['rooms'] = {'imported': len(rooms_list)}
                    _log.info(f'  Rooms saved OK')
                except Exception as e:
                    _log.error(f'  Rooms import error: {e}')
                    result['errors'].append(f'Failed to import Rooms: {e}')

            # Import Groups (case-insensitive)
            gname = _find_sheet_key('group')
            _log.info(f'  Groups sheet found as: {gname!r}')
            if gname:
                groups_stats = self._import_groups(wb[gname])
                result['groups'] = groups_stats
            else:
                result['errors'].append('Groups sheet not found')
            
            # Import Subjects (case-insensitive)
            sname = _find_sheet_key('subject')
            if sname:
                subjects_stats = self._import_subjects(wb[sname])
                result['subjects'] = subjects_stats
            else:
                result['errors'].append('Subjects sheet not found')
            
            # Import Group Schedules (case-insensitive)
            gs_name = _find_sheet_key('group schedule')
            imported_group_schedules = 0
            if gs_name:
                gs_ws = wb[gs_name]
                group_schedules = {}
                for row in gs_ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    g = row[0]
                    day = row[1]
                    lesson = int(row[2]) if row[2] is not None else None
                    subject = row[3] if len(row) > 3 else ''
                    teacher = row[4] if len(row) > 4 else ''
                    color_bg = row[5] if len(row) > 5 and row[5] else '#FFFFFF'
                    color_fg = row[6] if len(row) > 6 and row[6] else '#000000'
                    room = str(row[7]).strip() if len(row) > 7 and row[7] else ''

                    if g not in group_schedules:
                        group_schedules[g] = {}
                    if day not in group_schedules[g]:
                        group_schedules[g][day] = {}
                    if lesson:
                        group_schedules[g][day][int(lesson)] = {
                            'subject': subject or '',
                            'teacher': teacher or '',
                            'group': g,
                            'color_bg': color_bg,
                            'color_fg': color_fg,
                            'room': room,
                        }
                # Save each group's schedule
                for gname, sched in group_schedules.items():
                    try:
                        self.excel_service.save_group_schedule(gname, sched)
                        imported_group_schedules += 1
                    except Exception:
                        # continue on error
                        pass
                # After saving group schedules, rebuild teacher schedules
                try:
                    self.excel_service.rebuild_teacher_schedules()
                except Exception:
                    pass
                result['imported_group_schedules'] = imported_group_schedules

            # Import Teacher Schedules if present and group schedules not provided
            ts_name = _find_sheet_key('teacher schedule')
            imported_teacher_schedules = 0
            if ts_name and not gs_name:
                ts_ws = wb[ts_name]
                # Write directly into the ExcelService workbook Teacher Schedules sheet
                try:
                    if 'Teacher Schedules' not in self.excel_service.wb.sheetnames:
                        ws = self.excel_service.wb.create_sheet('Teacher Schedules')
                        ws.append(['Teacher', 'Day', 'Lesson', 'Subject', 'Group', 'Color BG', 'Color FG'])
                    else:
                        ws = self.excel_service.wb['Teacher Schedules']
                        # clear existing data
                        ws.delete_rows(2, ws.max_row)

                    for row in ts_ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        teacher = row[0]
                        day = row[1]
                        lesson = row[2]
                        subject = row[3] if len(row) > 3 else ''
                        group = row[4] if len(row) > 4 else ''
                        color_bg = row[5] if len(row) > 5 and row[5] else '#FFFFFF'
                        color_fg = row[6] if len(row) > 6 and row[6] else '#000000'
                        ws.append([teacher, day, lesson, subject, group, color_bg, color_fg])
                        imported_teacher_schedules += 1
                    result['imported_teacher_schedules'] = imported_teacher_schedules
                except Exception:
                    result['errors'].append('Failed to import Teacher Schedules')

            # Save after import
            try:
                self.excel_service.save()
            except Exception:
                pass
            
            _log.info(f'import_from_file END  result={result}')
            final_teachers = self.excel_service.get_teachers()
            _log.info(f'  FINAL get_teachers() count={len(final_teachers)} names={[t["name"] for t in final_teachers]}')
            _log.info('======================================================')
            return result
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'teachers': 0,
                'groups': 0,
                'subjects': 0
            }
    
    def _import_teachers(self, sheet):
        """Import teachers from Teachers sheet"""
        _log.info('=== _import_teachers START ===')
        _log.info(f'  Sheet title: {sheet.title!r}')
        _log.info(f'  max_row={sheet.max_row}  max_column={sheet.max_column}')
        # Log header row
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        _log.info(f'  HEADER row (col indices 0..{len(header)-1}): {header}')
        added = 0
        updated = 0
        existing_teachers = self.excel_service.get_teachers()
        _log.info(f'  Existing teachers in Excel BEFORE import: {[t["name"] for t in existing_teachers]}')
        teachers_list = list(existing_teachers)  # Convert to list for modification
        teacher_names_to_idx = {t['name']: i for i, t in enumerate(teachers_list)}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            teacher_name = row[0]
            subjects_raw = row[1] if len(row) > 1 else None
            total_hours = row[2] if len(row) > 2 else 0
            availability = row[3] if len(row) > 3 else None
            weekly_hours = row[4] if len(row) > 4 else None
            _log.debug(f'  ROW for teacher {teacher_name!r}: col0={row[0]!r} col1={str(subjects_raw)[:60]!r} col2={total_hours!r} col3={str(availability)[:40]!r} col4={str(weekly_hours)[:60]!r}')
            
            # Parse subjects (accept either newline-separated or semicolon-separated entries)
            subjects = []
            if subjects_raw:
                raw = str(subjects_raw)
                # split on newlines or semicolons (handles both 'a;b' and 'a; b' and multi-line)
                entries = re.split(r'(?:\r?\n|;)+', raw)
                for entry in entries:
                    line = entry.strip()
                    if not line:
                        continue
                    parts = [p.strip() for p in line.split(':')]
                    if len(parts) >= 3:
                        subjects.append({
                            'name': parts[0],
                            'hours': int(parts[1]) if parts[1].isdigit() else 0,
                            'group': parts[2]
                        })
            
            # Parse availability (format: "Day:Lesson-Lesson\nDay:Lesson-Lesson")
            availability_parsed = []
            if availability:
                for line in str(availability).split('\n'):
                    if ':' in line:
                        day, lessons = line.split(':', 1)
                        availability_parsed.append({
                            'day': day.strip(),
                            'lessons': lessons.strip()
                        })
            
            # Parse weekly hours (format: "Day:Time-Time\nDay:Time-Time")
            check_in_hours = {}
            check_out_hours = {}
            if weekly_hours:
                for line in str(weekly_hours).split('\n'):
                    line = line.strip()
                    if ':' in line and '-' in line:
                        # Split day and time parts
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            day = parts[0].strip()
                            time_part = parts[1].strip()
                            
                            # Extract time range
                            if '-' in time_part:
                                start_time, end_time = time_part.split('-', 1)
                                start_time = start_time.strip()
                                end_time = end_time.strip()
                                
                                # Add to dictionaries
                                if day not in check_in_hours:
                                    check_in_hours[day] = []
                                    check_out_hours[day] = []
                                
                                check_in_hours[day].append(start_time)
                                check_out_hours[day].append(end_time)
            
            # Create teacher data
            teacher_data = {
                'name': teacher_name,
                'subjects': subjects,
                'check_in_hours': check_in_hours,
                'check_out_hours': check_out_hours,
                'available_slots': {}
            }
            
            _log.debug(f'    parsed subjects count={len(subjects)}  check_in_days={list(check_in_hours.keys())}')
            if teacher_name in teacher_names_to_idx:
                # Update existing teacher
                idx = teacher_names_to_idx[teacher_name]
                teachers_list[idx] = teacher_data
                updated += 1
                _log.info(f'  UPDATED teacher: {teacher_name!r}')
            else:
                # Add new teacher
                teachers_list.append(teacher_data)
                added += 1
                _log.info(f'  ADDED teacher: {teacher_name!r}')
        
        _log.info(f'  Total teachers_list to save: {len(teachers_list)} — names: {[t["name"] for t in teachers_list]}')
        # Save all teachers
        self.excel_service.save_teachers(teachers_list)
        _log.info(f'  save_teachers() called. Verifying with get_teachers()...')
        verify = self.excel_service.get_teachers()
        _log.info(f'  AFTER save — get_teachers() returns {len(verify)} teachers: {[t["name"] for t in verify]}')
        if len(verify) != len(teachers_list):
            _log.error(f'  *** MISMATCH! sent {len(teachers_list)} but read back {len(verify)} ***')
        _log.info(f'=== _import_teachers END: added={added} updated={updated} ===')
        return {'added': added, 'updated': updated}
    
    def _import_groups(self, sheet):
        """Import groups from Groups sheet"""
        added = 0
        updated = 0
        existing_groups = self.excel_service.get_groups()
        groups_list = list(existing_groups)
        group_names_to_idx = {g['name']: i for i, g in enumerate(groups_list)}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            group_name = row[0]
            comment = row[1] if len(row) > 1 else None
            # Col 3: IsUnited (0/1 or blank)
            is_united_raw = row[2] if len(row) > 2 else None
            is_united = bool(is_united_raw) if is_united_raw is not None else False
            # Col 4: SubGroups (semicolon-separated)
            sub_groups_raw = row[3] if len(row) > 3 else None
            sub_groups = [s.strip() for s in str(sub_groups_raw).split(';') if s.strip()] if sub_groups_raw else []
            # Col 5: Room
            room = str(row[4]).strip() if len(row) > 4 and row[4] else ''

            group_data = {
                'name': group_name,
                'subjects': [str(comment)] if comment else [],
                'is_united': is_united,
                'sub_groups': sub_groups,
                'room': room,
            }
            
            if group_name in group_names_to_idx:
                idx = group_names_to_idx[group_name]
                # Preserve is_united/sub_groups from existing if not in import file
                existing = groups_list[idx]
                if not is_united and existing.get('is_united'):
                    group_data['is_united'] = existing['is_united']
                    group_data['sub_groups'] = existing.get('sub_groups', [])
                groups_list[idx] = group_data
                updated += 1
            else:
                groups_list.append(group_data)
                added += 1
        
        # Save all groups
        self.excel_service.save_groups(groups_list)
        return {'added': added, 'updated': updated}
    
    def _import_subjects(self, sheet):
        """Import subjects from Subjects sheet"""
        added = 0
        updated = 0
        existing_subjects = self.excel_service.get_subjects()
        subjects_list = list(existing_subjects)
        subject_keys_to_idx = {(s['name'], s['group']): i for i, s in enumerate(subjects_list)}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            subject_name = row[0]
            hours_per_week = row[1] if len(row) > 1 else 0
            group = row[2] if len(row) > 2 else ''
            
            # Parse hours_per_week if it's a string
            if isinstance(hours_per_week, str):
                hours_per_week = int(hours_per_week) if hours_per_week.isdigit() else 0
            
            subject_data = {
                'name': subject_name,
                'group': group,
                'hours_per_week': hours_per_week,
                'teacher': ''
            }
            
            key = (subject_name, group)
            if key in subject_keys_to_idx:
                idx = subject_keys_to_idx[key]
                subjects_list[idx] = subject_data
                updated += 1
            else:
                subjects_list.append(subject_data)
                added += 1
        
        # Save all subjects
        self.excel_service.save_subjects(subjects_list)
        return {'added': added, 'updated': updated}
