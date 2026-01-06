from openpyxl import load_workbook
import os

path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ExcelExamples', 'export_Good.xlsx')
if not os.path.exists(path):
    print('File not found:', path)
    exit(1)
wb = load_workbook(path, data_only=True)
print('Workbook:', os.path.basename(path))
print('Sheets:', wb.sheetnames)
print('---')
for s in wb.sheetnames:
    print('Sheet:', s)
    ws = wb[s]
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows[:5], start=1):
        print(i, r)
    print('... total rows:', len(rows))
    print('---')
