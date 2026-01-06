from openpyxl import load_workbook
import os

def dump_example_sheets(limit=15):
    """Return a list of tuples (sheet_name, sample_rows, total_rows).
    This function is non-printing so it can be used programmatically.
    """
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ExcelExamples', 'hours-2.xlsx')
    if not os.path.exists(path):
        return []

    wb = load_workbook(path, data_only=True)
    result = []
    for s in wb.sheetnames:
        if s.startswith('Group_') or s.startswith('Teacher_'):
            ws = wb[s]
            rows = list(ws.iter_rows(values_only=True))
            result.append((s, rows[:limit], len(rows)))
    return result

if __name__ == '__main__':
    # run silently
    dump_example_sheets()
