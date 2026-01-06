from openpyxl import load_workbook
import json
import glob
import os

def inspect_sheets(latest_only=True):
    uploads = 'uploads'
    export_files = sorted(glob.glob(os.path.join(uploads, 'export_*.xlsx')))
    latest_export = export_files[-1] if export_files else None
    paths = [os.path.join('ExcelExamples', 'hours-2.xlsx')]
    if latest_only and latest_export:
        paths.append(latest_export)

    result = {}
    for p in paths:
        try:
            wb = load_workbook(p, data_only=True)
        except Exception:
            continue
        sheets = {}
        for s in wb.sheetnames:
            ws = wb[s]
            first = next(ws.iter_rows(max_row=1, values_only=True), None)
            sheets[s] = first
        result[p] = sheets
    return result

if __name__ == '__main__':
    inspect_sheets()
