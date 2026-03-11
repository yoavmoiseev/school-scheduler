import openpyxl
wb = openpyxl.load_workbook('סיטרין ידני.xlsx')
out = open('read_output.txt', 'w', encoding='utf-8')

# Read schedules for all groups
target_sheets = [s for s in wb.sheetnames if s.startswith('Group_')]

for sheet_name in target_sheets:
    ws = wb[sheet_name]
    rows = [row for row in ws.iter_rows(values_only=True) if any(cell for cell in row)]
    if not rows:
        continue
    out.write(f'\n=== {sheet_name} ===\n')
    for row in rows:
        out.write(str(row) + '\n')

# Read teachers
out.write('\n\n=== TEACHERS ===\n')
ws_teachers = wb['Teachers']
for row in ws_teachers.iter_rows(values_only=True):
    if any(cell for cell in row):
        out.write(str(row) + '\n')

out.close()
print('Done')
