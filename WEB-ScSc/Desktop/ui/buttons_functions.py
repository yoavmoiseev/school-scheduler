import os
import shutil
from datetime import datetime
from tkinter import messagebox
from school_scheduler.data.consts import DATA_DIR, SAVED_SCHEDULES_DIR
from school_scheduler.ui.schedule_io import reload_all
import time
import tkinter as tk
import pyautogui
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from PIL import Image
from reportlab.lib.utils import ImageReader
from . import translations as tr




# ===================================================================================================================
from tkinter import filedialog
def save_as_snapshot() -> str:
    data_dir = os.path.abspath(DATA_DIR)

    if not os.path.exists(data_dir):
        messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[50]} {data_dir}")  # ❌ Папка DATA_DIR не найдена
        return ""

    # Выбор папки через диалог
    target_dir = filedialog.askdirectory(title=tr.GUI_TEXT[53])  # вместо .get()
    if not target_dir:
        return ""  # пользователь отменил

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_folder = os.path.join(target_dir, f"backup_{timestamp}")

    try:
        shutil.copytree(data_dir, new_folder)
        messagebox.showinfo(tr.GUI_TEXT[70] if len(tr.GUI_TEXT) > 70 else "Info", f"{tr.GUI_TEXT[51]} {new_folder}")  # ✅ Папка data сохранена как
        return new_folder
    except Exception as e:
        messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[52]} {e}")  # ❌ Ошибка при сохранении
        return ""

# ===================================================================================================================
def load_snapshot():

    
    data_dir = os.path.abspath(DATA_DIR)

    # Выбор папки для загрузки
    source_dir = filedialog.askdirectory(title=tr.GUI_TEXT[54])  # вместо .get()
    if not source_dir:
        return  # пользователь отменил

    if not os.path.exists(source_dir):
        messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[55]} {source_dir}")  # ❌ Папка для загрузки не найдена
        return

    # Verify required JSON files exist in the selected folder
    required_files = ["groups.json", "subjects.json", "teachers.json", "schedules.json"]
    missing = [f for f in required_files if not os.path.exists(os.path.join(source_dir, f))]
    if missing:
        # Show which files are missing and abort (localized)
        try:
            title = tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error"
            messagebox.showerror(title, f"{tr.GUI_TEXT[124]} {', '.join(missing)}") # ❌ The selected folder is missing required files:
        except Exception:
            # fallback
            try:
                messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[124]} {', '.join(missing)}")
            except Exception:
                pass
        return

    # Удаляем старое содержимое DATA_DIR
    if os.path.exists(data_dir):
        try:
            shutil.rmtree(data_dir)
        except Exception as e:
            messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[56]} {e}")  # ❌ Ошибка при очистке DATA_DIR
            return

    # Копируем новую папку
    try:
        shutil.copytree(source_dir, data_dir)
        messagebox.showinfo(tr.GUI_TEXT[70] if len(tr.GUI_TEXT) > 70 else "Info", f"{tr.GUI_TEXT[125]} {data_dir}")  # ✅ Данные успешно загружены
    except Exception as e:
        messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", f"{tr.GUI_TEXT[52]} {e}")  # ❌ Ошибка при загрузке
        return

    # Полный перезапуск приложения
    import sys
    import subprocess
    
    # Закрываем текущее окно
    root = tk._get_default_root()
    if root:
        root.quit()
        root.destroy()
    
    # Запускаем новый процесс приложения
    env = os.environ.copy()
    if getattr(sys, 'frozen', False):
        # Running as EXE: sys.executable is the EXE path
        subprocess.Popen([sys.executable] + sys.argv[1:], env=env)
    else:
        # Running as script: need python + script path
        subprocess.Popen([sys.executable] + sys.argv, env=env)
    sys.exit()
    


# ===================================================================================================================



def rebuild_all_schedule(app, reverse=False, randomize=False, praiority_groups=None):
    """
    Deletes the existing schedules.json file and rebuilds the schedule for all groups.
    """

    # "Rebuild the schedule" / "Delete and rebuild last schedule?"
    choice = messagebox.askokcancel(tr.GUI_TEXT[53], tr.GUI_TEXT[54], default=messagebox.CANCEL)
    if not choice:
        return

    # Backup current schedule data
    save_as_snapshot()

    # delete school_scheduler\data\schedules.json file
    schedules_file = os.path.join(DATA_DIR, "schedules.json")
    if os.path.exists(schedules_file):
        os.remove(schedules_file)

    # Reload all data and rebuild schedules for all groups
    # reload_all(app) # BUG why reload requiered???

    # Determine group order
    if praiority_groups is not None:
        groups_list = praiority_groups
    else:
        if reverse:
            groups_list = app.groups[::-1]
        else:
            groups_list = app.groups
        if randomize:
            import random
            random.shuffle(groups_list)

    app.group_var.set(groups_list[0].name if groups_list else "")
    for group in groups_list:
        app.clear_cells(app.group_cells)
        app.autofill_group(group.name, app.group_cells, app.subjects, app.teachers)
        app.save_group_schedule(group.name, app.group_cells, user_confirmation=False)

    # "Расписание пересоздано."
    messagebox.showinfo(tr.GUI_TEXT[70] if len(tr.GUI_TEXT) > 70 else "Info", tr.GUI_TEXT[55])  # original: "Расписание пересоздано."


def print_to_pdf(frame=None, name=None, output_path=None):
    """
    Takes a screenshot of the specified Tkinter frame (or the whole window)
    and saves it into a Landscape-oriented PDF file.
    Works stably even with multiple monitors.
    """

    # Determine coordinates
    if frame:
        x = frame.winfo_rootx()
        y = frame.winfo_rooty()
        w = frame.winfo_width()
        h = frame.winfo_height()
    else:
        root = tk._get_default_root()
        if not root:
            # "Ошибка: Нет активного окна для скриншота."
            messagebox.showerror(tr.GUI_TEXT[67] if len(tr.GUI_TEXT) > 67 else "Error", tr.GUI_TEXT[56])  # original: "Ошибка: Нет активного окна для скриншота."
            return
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        w = root.winfo_width()
        h = root.winfo_height()

    # Take screenshot
    screenshot = pyautogui.screenshot(region=(x, y, w, h))

    # Define output path
    if not output_path:
        os.makedirs("exports", exist_ok=True)
        output_path = os.path.join(
            "exports",
            f"{name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
        )

    # Save temporary PNG
    temp_img_path = output_path.replace(".pdf", ".png")
    screenshot.save(temp_img_path)

    # Create Landscape PDF
    c = canvas.Canvas(output_path, pagesize=landscape(A4))
    pdf_width, pdf_height = landscape(A4)

    with Image.open(temp_img_path) as img:
        img_width, img_height = img.size
        ratio = min(pdf_width / img_width, pdf_height / img_height)
        new_width = img_width * ratio
        new_height = img_height * ratio
        x_pos = (pdf_width - new_width) / 2
        y_pos = (pdf_height - new_height) / 2
        c.drawImage(ImageReader(img), x_pos, y_pos, new_width, new_height)

    c.showPage()
    c.save()

    # Delete temp PNG
    try:
        os.remove(temp_img_path)
    except PermissionError:
        pass

    # ✅ PDF сохранён (Landscape)
    messagebox.showinfo(tr.GUI_TEXT[70] if len(tr.GUI_TEXT) > 70 else "Info", f"{tr.GUI_TEXT[57]}\n{output_path}")  # original: "✅ PDF сохранён (Landscape)"


# ===================================================================================================================
# Clear functions
# ===================================================================================================================

def clear_schedulers(app):
    """Clear only schedules.json, keep Teachers, Groups, Subjects"""
    # Show warning dialog with Cancel as default
    choice = messagebox.askokcancel(
        tr.GUI_TEXT[141],  # "Clear Schedulers"
        tr.GUI_TEXT[143],  # Warning message
        default=messagebox.CANCEL
    )
    
    if not choice:
        return
    
    # Create backup before clearing
    save_as_snapshot()
    
    # Clear schedules.json
    schedules_file = os.path.join(DATA_DIR, "schedules.json")
    if os.path.exists(schedules_file):
        try:
            os.remove(schedules_file)
        except Exception as e:
            messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[194].format(error=str(e)))  # "Error", "Error deleting schedules.json: {error}"
            return
    
    # Reload UI
    reload_all(app)
    
    # Show success message
    messagebox.showinfo(tr.GUI_TEXT[70], tr.GUI_TEXT[145])  # "Schedules cleared successfully."


def clear_all_data(app):
    """Clear ALL data: teachers.json, groups.json, subjects.json, schedules.json"""
    # Show warning dialog with Cancel as default
    choice = messagebox.askokcancel(
        tr.GUI_TEXT[142],  # "Clear All Data"
        tr.GUI_TEXT[144],  # Warning message
        default=messagebox.CANCEL
    )
    
    if not choice:
        return
    
    # Create backup before clearing
    save_as_snapshot()
    
    # List of all JSON files to clear
    files_to_clear = [
        "teachers.json",
        "groups.json",
        "subjects.json",
        "schedules.json"
    ]
    
    # Delete all JSON files
    for filename in files_to_clear:
        file_path = os.path.join(DATA_DIR, filename)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[195].format(filename=filename, error=str(e)))  # "Error", "Error deleting {filename}: {error}"
                return
    
    # Reload UI
    reload_all(app)
    
    # Show success message
    messagebox.showinfo(tr.GUI_TEXT[70], tr.GUI_TEXT[146])  # "All data cleared successfully."


# ===================================================================================================================
# Excel Export/Import functions
# ===================================================================================================================

def export_to_excel(app):
    """Export all data (Teachers, Groups, Subjects, Schedules) to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[180])  # "Error", "openpyxl library is not installed..."
        return
    
    # Ask where to save
    file_path = filedialog.asksaveasfilename(
        title=tr.GUI_TEXT[147],  # "Export to Excel"
        defaultextension=".xlsx",
        filetypes=[(tr.GUI_TEXT[197], "*.xlsx"), (tr.GUI_TEXT[198], "*.*")]  # "Excel files", "All files"
    )
    
    if not file_path:
        return
    
    try:
        import school_scheduler.data.consts as consts
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Style definitions
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # ========== TEACHERS SHEET ==========
        ws_teachers = wb.create_sheet(title="Teachers")
        ws_teachers['A1'] = "Teacher Name"
        ws_teachers['B1'] = "Subjects (Name:Hours:Group)"
        ws_teachers['C1'] = "Total Hours"
        ws_teachers['D1'] = "Availability (Day:Start-End)"
        ws_teachers['E1'] = "Teachers Weekly Hours (Day:Start-End)"
        
        # Apply header style
        for cell in ws_teachers[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = cell_border
        
        # Data rows
        for row_idx, teacher in enumerate(app.teachers, start=2):
            ws_teachers.cell(row_idx, 1, teacher.name).border = cell_border
            
            # Subjects
            subjects_str = "\n".join([f"{s['name']}:{s.get('hours', 0)}:{s.get('group', '')}" for s in teacher.subjects])
            cell = ws_teachers.cell(row_idx, 2, subjects_str)
            cell.border = cell_border
            cell.alignment = wrap_align
            
            # Total hours
            total_hours = sum([s.get('hours', 0) for s in teacher.subjects])
            ws_teachers.cell(row_idx, 3, total_hours).border = cell_border
            
            # Availability
            avail_slots = getattr(teacher, 'available_slots', {})
            avail_str = "\n".join([f"{day}:{start}-{end}" for day, intervals in avail_slots.items() for start, end in intervals])
            cell = ws_teachers.cell(row_idx, 4, avail_str)
            cell.border = cell_border
            cell.alignment = wrap_align
            
            # Teachers Weekly Hours (combine check_in and check_out into Start-End format)
            check_in_hours = getattr(teacher, 'check_in_hours', {})
            check_out_hours = getattr(teacher, 'check_out_hours', {})
            
            # Build weekly hours list: "Day:Start-End" for each interval
            weekly_hours_list = []
            for day, intervals in avail_slots.items():
                # Get check_in/check_out for this day (can be string or list)
                day_check_ins = check_in_hours.get(day, [])
                day_check_outs = check_out_hours.get(day, [])
                
                # Ensure they are lists
                if isinstance(day_check_ins, str):
                    day_check_ins = [day_check_ins] if day_check_ins else []
                if isinstance(day_check_outs, str):
                    day_check_outs = [day_check_outs] if day_check_outs else []
                
                # Add one line for each interval
                for idx, (start_lesson, end_lesson) in enumerate(intervals):
                    check_in = day_check_ins[idx] if idx < len(day_check_ins) else ""
                    check_out = day_check_outs[idx] if idx < len(day_check_outs) else ""
                    
                    # If no check_in/check_out, calculate from TIME_SLOTS (like GUI does)
                    if not check_in or not check_out:
                        time_start = consts.TIME_SLOTS.get(start_lesson, "").split("-")[0]
                        time_end = consts.TIME_SLOTS.get(end_lesson, "").split("-")[1] if "-" in consts.TIME_SLOTS.get(end_lesson, "") else ""
                        if time_start and time_end:
                            check_in = time_start
                            check_out = time_end
                    
                    if check_in and check_out:
                        weekly_hours_list.append(f"{day}:{check_in}-{check_out}")
            
            weekly_hours_str = "\n".join(weekly_hours_list)
            cell = ws_teachers.cell(row_idx, 5, weekly_hours_str)
            cell.border = cell_border
            cell.alignment = wrap_align
        
        # Adjust column widths
        ws_teachers.column_dimensions['A'].width = 20
        ws_teachers.column_dimensions['B'].width = 40
        ws_teachers.column_dimensions['C'].width = 12
        ws_teachers.column_dimensions['D'].width = 30
        ws_teachers.column_dimensions['E'].width = 30
        
        # ========== GROUPS SHEET ==========
        ws_groups = wb.create_sheet(title="Groups")
        ws_groups['A1'] = "Group Name"
        ws_groups['B1'] = "Comment"
        
        # Apply header style
        for cell in ws_groups[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = cell_border
        
        # Data rows
        for row_idx, group in enumerate(app.groups, start=2):
            ws_groups.cell(row_idx, 1, group.name).border = cell_border
            ws_groups.cell(row_idx, 2, group.comment).border = cell_border
        
        # Adjust column widths
        ws_groups.column_dimensions['A'].width = 20
        ws_groups.column_dimensions['B'].width = 30
        
        # ========== SUBJECTS SHEET ==========
        ws_subjects = wb.create_sheet(title="Subjects")
        ws_subjects['A1'] = "Subject Name"
        ws_subjects['B1'] = "Hours per Week"
        ws_subjects['C1'] = "Group"
        
        # Apply header style
        for cell in ws_subjects[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = cell_border
        
        # Data rows
        for row_idx, subject in enumerate(app.subjects, start=2):
            ws_subjects.cell(row_idx, 1, subject.name).border = cell_border
            ws_subjects.cell(row_idx, 2, subject.hours_per_week).border = cell_border
            group_name = getattr(subject, 'group', '') or ''
            ws_subjects.cell(row_idx, 3, group_name).border = cell_border
        
        # Adjust column widths
        ws_subjects.column_dimensions['A'].width = 25
        ws_subjects.column_dimensions['B'].width = 15
        ws_subjects.column_dimensions['C'].width = 20
        
        # Export Group Schedules
        for group in app.groups:
            ws = wb.create_sheet(title=f"Group_{group.name}")
            schedule_data = app.schedules.get('groups', {}).get(group.name, {})
            
            # Headers
            ws['A1'] = "Lesson"
            for col_idx, day in enumerate(app.schedules.get('weekdays', consts.WEEKDAYS), start=2):
                ws.cell(1, col_idx, day)
            
            # Apply header style
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = cell_border
            
            # Data rows
            for lesson_num in sorted(consts.TIME_SLOTS.keys()):
                row_idx = lesson_num + 1
                time_slot = consts.TIME_SLOTS.get(lesson_num, "")
                ws.cell(row_idx, 1, f"{lesson_num} ({time_slot})")
                ws.cell(row_idx, 1).border = cell_border
                ws.cell(row_idx, 1).alignment = center_align
                
                lesson_str = str(lesson_num)
                day_map = schedule_data.get(lesson_str, {})
                
                for col_idx, day in enumerate(app.schedules.get('weekdays', consts.WEEKDAYS), start=2):
                    subject = day_map.get(day, "")
                    cell = ws.cell(row_idx, col_idx, subject)
                    cell.border = cell_border
                    cell.alignment = center_align
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            for col in range(2, 7):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Export Teacher Schedules
        for teacher in app.teachers:
            ws = wb.create_sheet(title=f"Teacher_{teacher.name}"[:31])  # Excel limit 31 chars
            schedule_data = app.schedules.get('teachers', {}).get(teacher.name, {})
            
            # Headers
            ws['A1'] = "Lesson"
            for col_idx, day in enumerate(app.schedules.get('weekdays', consts.WEEKDAYS), start=2):
                ws.cell(1, col_idx, day)
            
            # Apply header style
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = cell_border
            
            # Data rows
            for lesson_num in sorted(consts.TIME_SLOTS.keys()):
                row_idx = lesson_num + 1
                time_slot = consts.TIME_SLOTS.get(lesson_num, "")
                ws.cell(row_idx, 1, f"{lesson_num} ({time_slot})")
                ws.cell(row_idx, 1).border = cell_border
                ws.cell(row_idx, 1).alignment = center_align
                
                lesson_str = str(lesson_num)
                day_map = schedule_data.get(lesson_str, {})
                
                for col_idx, day in enumerate(app.schedules.get('weekdays', consts.WEEKDAYS), start=2):
                    subject = day_map.get(day, "")
                    cell = ws.cell(row_idx, col_idx, subject)
                    cell.border = cell_border
                    cell.alignment = center_align
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            for col in range(2, 7):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(file_path)
        messagebox.showinfo(tr.GUI_TEXT[181], tr.GUI_TEXT[182].format(file_path=file_path))  # "Success", "Schedules exported successfully to:..."
        
    except Exception as e:
        messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[183].format(error=str(e)))  # "Error", "Failed to export to Excel:..."


def load_from_excel(app):
    """Load Teachers, Groups, Subjects and Schedules from Excel file"""
    try:
        import openpyxl
    except ImportError:
        messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[180])  # "Error", "openpyxl library is not installed..."
        return
    
    # Ask which file to load
    file_path = filedialog.askopenfilename(
        title=tr.GUI_TEXT[148],  # "Load from Excel"
        filetypes=[(tr.GUI_TEXT[197], "*.xlsx"), (tr.GUI_TEXT[198], "*.*")]  # "Excel files", "All files"
    )
    
    if not file_path:
        return
    
    try:
        wb = openpyxl.load_workbook(file_path)
        from school_scheduler.helpers import json_io
        import school_scheduler.data.consts as consts
        from models.teacher import Teacher
        from models.group import Group
        from models.subject import Subject
        import os
        
        loaded_items = []
        
        # ========== IMPORT TEACHERS ==========
        if "Teachers" in wb.sheetnames:
            ws = wb["Teachers"]
            teachers_data = []
            
            for row_idx in range(2, ws.max_row + 1):
                name = ws.cell(row_idx, 1).value
                if not name:
                    continue
                
                # Parse subjects
                subjects_str = ws.cell(row_idx, 2).value or ""
                subjects = []
                for subj_line in str(subjects_str).split('\n'):
                    if ':' in subj_line:
                        parts = subj_line.split(':')
                        if len(parts) >= 2:
                            subj_name = parts[0].strip()
                            try:
                                hours = int(parts[1].strip())
                            except:
                                hours = 0
                            group = parts[2].strip() if len(parts) >= 3 else ""
                            subjects.append({"name": subj_name, "hours": hours, "group": group})
                
                # Parse availability
                avail_str = ws.cell(row_idx, 4).value or ""
                available_slots = {}
                for avail_line in str(avail_str).split('\n'):
                    if ':' in avail_line and '-' in avail_line:
                        parts = avail_line.split(':')
                        day = parts[0].strip()
                        range_parts = parts[1].split('-')
                        try:
                            start = int(range_parts[0].strip())
                            end = int(range_parts[1].strip())
                            available_slots.setdefault(day, []).append((start, end))
                        except:
                            pass
                
                # Parse Teachers Weekly Hours (column E) - format: "Day:Start-End"
                # Can have multiple lines for same day (different intervals)
                weekly_hours_str = ws.cell(row_idx, 5).value or ""
                check_in_hours = {}
                check_out_hours = {}
                for line in str(weekly_hours_str).split('\n'):
                    if ':' in line and '-' in line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            day = parts[0].strip()
                            time_range = parts[1].strip()
                            if day and time_range and '-' in time_range:
                                time_parts = time_range.split('-', 1)
                                if len(time_parts) == 2:
                                    # Store as lists to support multiple intervals per day
                                    if day not in check_in_hours:
                                        check_in_hours[day] = []
                                    if day not in check_out_hours:
                                        check_out_hours[day] = []
                                    check_in_hours[day].append(time_parts[0].strip())
                                    check_out_hours[day].append(time_parts[1].strip())
                
                # PRIORITY: Always use Teachers Weekly Hours if present
                # Auto-generate available_slots from check_in/check_out times
                if check_in_hours and check_out_hours:
                    # Clear existing available_slots, will rebuild from Teachers Weekly Hours
                    available_slots = {}
                    
                    for day in check_in_hours:
                        if day in check_out_hours:
                            # Get lists of times for this day
                            day_check_ins = check_in_hours[day] if isinstance(check_in_hours[day], list) else [check_in_hours[day]]
                            day_check_outs = check_out_hours[day] if isinstance(check_out_hours[day], list) else [check_out_hours[day]]
                            
                            # Process each time pair
                            for idx in range(min(len(day_check_ins), len(day_check_outs))):
                                checkin_str = day_check_ins[idx]
                                checkout_str = day_check_outs[idx]
                                
                                # Convert check in/out times to lesson numbers
                                try:
                                    checkin_h, checkin_m = map(int, checkin_str.split(":"))
                                    checkout_h, checkout_m = map(int, checkout_str.split(":"))
                                    checkin_minutes = checkin_h * 60 + checkin_m
                                    checkout_minutes = checkout_h * 60 + checkout_m
                                    
                                    if checkout_minutes > checkin_minutes:
                                        first_lesson = None
                                        last_lesson = None
                                        
                                        # Find which lessons are covered by this time range
                                        for lesson_num in sorted(consts.TIME_SLOTS.keys()):
                                            time_range = consts.TIME_SLOTS[lesson_num]
                                            if "-" not in time_range:
                                                continue
                                            
                                            lesson_start, lesson_end = time_range.split("-")
                                            ls_h, ls_m = map(int, lesson_start.split(":"))
                                            le_h, le_m = map(int, lesson_end.split(":"))
                                            lesson_start_minutes = ls_h * 60 + ls_m
                                            lesson_end_minutes = le_h * 60 + le_m
                                            
                                            # Check if lesson is fully covered by check in/out time
                                            if checkin_minutes <= lesson_start_minutes and checkout_minutes >= lesson_end_minutes:
                                                if first_lesson is None:
                                                    first_lesson = lesson_num
                                                last_lesson = lesson_num
                                        
                                        # Add to available_slots if we found lessons
                                        if first_lesson is not None and last_lesson is not None:
                                            available_slots.setdefault(day, []).append((first_lesson, last_lesson))
                                except (ValueError, AttributeError):
                                    pass  # Skip invalid time formats
                
                teachers_data.append({
                    "name": name,
                    "subjects": subjects,
                    "available_slots": available_slots,
                    "check_in_hours": check_in_hours,
                    "check_out_hours": check_out_hours
                })
            
            if teachers_data:
                json_io.write_save_json(os.path.join(consts.DATA_DIR, "teachers.json"), "teachers", teachers_data)
                loaded_items.append(f"Teachers ({len(teachers_data)})")
        
        # ========== IMPORT GROUPS ==========
        if "Groups" in wb.sheetnames:
            ws = wb["Groups"]
            groups_data = []
            
            for row_idx in range(2, ws.max_row + 1):
                name = ws.cell(row_idx, 1).value
                comment = ws.cell(row_idx, 2).value
                
                if not name:
                    continue
                
                comment = str(comment).strip() if comment else ""
                
                groups_data.append({"name": name, "comment": comment})
            
            if groups_data:
                json_io.write_save_json(os.path.join(consts.DATA_DIR, "groups.json"), "groups", groups_data)
                loaded_items.append(f"Groups ({len(groups_data)})")
        
        # ========== IMPORT SUBJECTS ==========
        if "Subjects" in wb.sheetnames:
            ws = wb["Subjects"]
            subjects_data = []
            
            for row_idx in range(2, ws.max_row + 1):
                name = ws.cell(row_idx, 1).value
                hours = ws.cell(row_idx, 2).value
                group = ws.cell(row_idx, 3).value
                
                if not name:
                    continue
                
                try:
                    hours = int(hours) if hours else 0
                except:
                    hours = 0
                
                subjects_data.append({
                    "name": name,
                    "hours_per_week": hours,
                    "group": group or ""
                })
            
            if subjects_data:
                json_io.write_save_json(os.path.join(consts.DATA_DIR, "subjects.json"), "subjects", subjects_data)
                loaded_items.append(f"Subjects ({len(subjects_data)})")
        
        # ========== IMPORT GROUP SCHEDULES ==========
        imported_schedules = 0
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Group_"):
                group_name = sheet_name.replace("Group_", "")
                
                ws = wb[sheet_name]
                schedule_data = {}
                
                # Read headers (days) from row 1
                days = []
                for col_idx in range(2, ws.max_column + 1):
                    day = ws.cell(1, col_idx).value
                    if day:
                        days.append((col_idx, day))
                
                # Read data rows
                for row_idx in range(2, ws.max_row + 1):
                    lesson_cell = ws.cell(row_idx, 1).value
                    if not lesson_cell:
                        continue
                    
                    # Extract lesson number
                    lesson_str = str(lesson_cell).split()[0] if ' ' in str(lesson_cell) else str(lesson_cell)
                    
                    try:
                        lesson_num = int(lesson_str)
                    except ValueError:
                        continue
                    
                    lesson_key = str(lesson_num)
                    day_map = {}
                    
                    for col_idx, day in days:
                        subject = ws.cell(row_idx, col_idx).value
                        if subject and str(subject).strip():
                            day_map[day] = str(subject).strip()
                    
                    if day_map:
                        schedule_data[lesson_key] = day_map
                
                if schedule_data:
                    app.schedules.setdefault('groups', {})[group_name] = schedule_data
                    imported_schedules += 1
        
        if imported_schedules > 0:
            json_io.write_save_dict(consts.SCHEDULES_FILE, app.schedules)
            loaded_items.append(f"Schedules ({imported_schedules})")
        
        if loaded_items:
            # Reload application data
            app.reload_all()
            messagebox.showinfo(tr.GUI_TEXT[181], tr.GUI_TEXT[184].format(items="\n".join(loaded_items)))  # "Success", "Imported from Excel:..."
        else:
            messagebox.showwarning(tr.GUI_TEXT[157], tr.GUI_TEXT[185])  # "Warning", "No data found in Excel file"
        
    except Exception as e:
        messagebox.showerror(tr.GUI_TEXT[67], tr.GUI_TEXT[186].format(error=str(e)))  # "Error", "Failed to load from Excel:..."

