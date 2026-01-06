from openpyxl import load_workbook
import glob, os

def analyze_examples(limit_samples=10):
    """Analyze example workbooks for non-empty Group_/Teacher_ sheets and return a dict.
    Returns: {filename: {sheet_name: {'non_empty': int, 'samples': [(row_index, value), ...]}}}
    """
    base = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ExcelExamples')
    files = glob.glob(os.path.join(base, '*.xlsx'))
    if not files:
        return {}

    report = {}
    for f in files:
        try:
            wb = load_workbook(f, data_only=True)
        except Exception:
            continue
        file_report = {}
        for s in wb.sheetnames:
            if s.startswith('Group_') or s.startswith('Teacher_'):
                ws = wb[s]
                non_empty = 0
                sample = []
                rows = list(ws.iter_rows(values_only=True))
                for idx, r in enumerate(rows[1:], start=2):
                    for c in r[1:]:
                        if c is not None and str(c).strip() != '':
                            non_empty += 1
                            if len(sample) < limit_samples:
                                sample.append((idx, c))
                file_report[s] = {'non_empty': non_empty, 'samples': sample}
        if file_report:
            report[os.path.basename(f)] = file_report
    return report

if __name__ == '__main__':
    analyze_examples()
