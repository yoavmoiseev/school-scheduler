from flask import Blueprint, request, jsonify
import os
import shutil
from datetime import datetime

admin_bp = Blueprint('admin', __name__)

# Will be initialized by app.py
excel_service = None
autofill_service = None


def init_routes(excel_svc, autofill_svc=None):
    global excel_service, autofill_service
    excel_service = excel_svc
    autofill_service = autofill_svc


def _detect_data_dir():
    """Detect the most likely data directory used by the app/desktop code."""
    # 1. Try to import desktop consts if available
    try:
        import school_scheduler.data.consts as ds
        dd = getattr(ds, 'DATA_DIR', None)
        if dd and os.path.exists(dd):
            return dd
    except Exception:
        pass

    # 2. Try common Desktop/data folder relative to project
    here = os.path.dirname(os.path.dirname(__file__))
    cand = os.path.join(here, 'Desktop', 'data')
    if os.path.exists(cand):
        return cand

    # 3. Fallback to app data folder (relative cwd)
    return os.path.abspath('data')


@admin_bp.route('/api/admin/backups', methods=['GET'])
def list_backups():
    uploads = os.path.abspath('uploads')
    backups_dir = os.path.join(uploads, 'backups')
    os.makedirs(backups_dir, exist_ok=True)
    items = []
    for name in os.listdir(backups_dir):
        path = os.path.join(backups_dir, name)
        if os.path.isdir(path):
            items.append(name)
    return jsonify({'success': True, 'backups': items})


@admin_bp.route('/api/admin/backup', methods=['POST'])
def create_backup():
    data_dir = os.path.abspath('data')
    if not os.path.exists(data_dir):
        return jsonify({'success': False, 'error': 'Data folder not found'}), 404

    uploads = os.path.abspath('uploads')
    backups_dir = os.path.join(uploads, 'backups')
    os.makedirs(backups_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(backups_dir, f'backup_{timestamp}')
    try:
        shutil.copytree(data_dir, dest)
        return jsonify({'success': True, 'backup': os.path.basename(dest)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/admin/restore', methods=['POST'])
def restore_backup():
    body = request.json or {}
    backup_name = body.get('backup')
    if not backup_name:
        return jsonify({'success': False, 'error': 'backup name required'}), 400
    uploads = os.path.abspath('uploads')
    backups_dir = os.path.join(uploads, 'backups')
    source = os.path.join(backups_dir, backup_name)
    data_dir = os.path.abspath('data')
    if not os.path.exists(source):
        return jsonify({'success': False, 'error': 'backup not found'}), 404

    # remove existing data dir
    try:
        if os.path.exists(data_dir):
            shutil.rmtree(data_dir)
        shutil.copytree(source, data_dir)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/admin/clear-all', methods=['POST'])
def clear_all_data():
    # Prefer clearing via ExcelService if available
    try:
        if excel_service:
            # Clear teachers, groups, subjects and schedules inside Excel
            try:
                excel_service.save_teachers([])
            except Exception:
                pass
            try:
                excel_service.save_groups([])
            except Exception:
                pass
            try:
                excel_service.save_subjects([])
            except Exception:
                pass
            try:
                excel_service.clear_all_schedules()
            except Exception:
                pass

            return jsonify({'success': True, 'cleared': 'excel'})

        # Fallback to removing JSON files if ExcelService not available
        data_dir = _detect_data_dir()
        files = ['teachers.json', 'groups.json', 'subjects.json', 'schedules.json']
        removed = []
        missing = []
        for f in files:
            p = os.path.join(data_dir, f)
            if os.path.exists(p):
                os.remove(p)
                removed.append(f)
            else:
                missing.append(f)
        return jsonify({'success': True, 'data_dir': data_dir, 'removed': removed, 'missing': missing})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/admin/rebuild-all', methods=['POST'])
def rebuild_all():
    # Rebuild schedules for all groups using autofill_service if available
    try:
        # quick debug log to trace rebuild requests
        try:
            os.makedirs('uploads', exist_ok=True)
            with open(os.path.join('uploads', 'rebuild_all.log'), 'a', encoding='utf-8') as lf:
                lf.write(f"{datetime.now().isoformat()} | rebuild_all called | body={str(request.json or {})}\n")
        except Exception:
            pass
        body = request.json or {}
        priorities = body.get('priorities') or {}

        # Load configuration (used for max_autofill_retries fallback)
        config = excel_service.get_config() if excel_service else {}
        try:
            max_autofill_retries = int(config.get('max_autofill_retries', 100))
        except Exception:
            max_autofill_retries = 100

        # Ensure ExcelService reloads the workbook to pick up any recent external changes
        try:
            if excel_service:
                excel_service.load()
        except Exception:
            pass
        groups = excel_service.get_groups() if excel_service else []
        # If Excel has no groups, try to read groups.json from detected data dir
        if (not groups or len(groups) == 0):
            data_dir = _detect_data_dir()
            groups_path = os.path.join(data_dir, 'groups.json')
            if os.path.exists(groups_path):
                import json
                try:
                    with open(groups_path, 'r', encoding='utf-8') as fh:
                        data = json.load(fh)
                        groups = data.get('groups') or data.get('items') or []
                except Exception:
                    groups = []
        # If autofill_service is available, use it; otherwise call schedules save via existing logic is limited
        if autofill_service:
            # If priorities.groups provided, use that order (strings); otherwise use groups list
            ordered = []
            if priorities and isinstance(priorities, dict) and priorities.get('groups'):
                # priorities groups list may contain names
                ordered = priorities.get('groups')
            else:
                # normalize from groups list — skip united groups (virtual, no direct schedule)
                for g in groups:
                    name = g.get('name') if isinstance(g, dict) else (g.get('name') if hasattr(g, 'get') else None)
                    if not name:
                        name = g if isinstance(g, str) else None
                    if name:
                        ordered.append(name)

            # Build a map of group name -> group object for quick lookup
            groups_map = {g.get('name'): g for g in groups if isinstance(g, dict) and g.get('name')}

            # Separate united groups from regular groups.
            # United groups must be processed FIRST so their slots are marked busy
            # before regular sub-groups are filled. After autofill, their lessons
            # are propagated into each sub-group's schedule.
            united_names_ordered = []
            regular_names_ordered = []
            for n in ordered:
                g = groups_map.get(n)
                if g and g.get('is_united'):
                    united_names_ordered.append(n)
                else:
                    regular_names_ordered.append(n)

            # Also include any united groups not already in ordered
            all_united = [g.get('name') for g in groups if isinstance(g, dict) and g.get('is_united') and g.get('name')]
            for n in all_united:
                if n not in united_names_ordered:
                    united_names_ordered.append(n)

            # ── Phase 1: clear all schedules so we start from scratch ──────────
            try:
                excel_service.clear_all_schedules()
            except Exception:
                pass

            # Extract subject order from priorities (list of subject name strings)
            subject_order = priorities.get('subjects', []) if priorities and isinstance(priorities, dict) else []

            rebuild_log = []

            # ── Phase 2: build united groups (fresh) ─────────────────────────────
            # After each united group is built, force the resulting lessons
            # into every sub-group so those slots are reserved before the
            # sub-group's own autofill runs.
            for name in united_names_ordered:
                if not name:
                    continue
                try:
                    try:
                        if excel_service:
                            excel_service.load()
                    except Exception:
                        pass
                    outer_max = int(config.get('max_autofill_retries', max_autofill_retries)) if config else max_autofill_retries
                    attempt = 0
                    success = False
                    last_info = None
                    while attempt < outer_max and not success:
                        try:
                            success, schedule, info = autofill_service.autofill_group(name, max_retries=1, preserve_existing=False, subject_order=subject_order)
                        except TypeError:
                            success, schedule, info = autofill_service.autofill_group(name)
                        last_info = info
                        if schedule and any(schedule.values()):
                            try:
                                excel_service.save_group_schedule(name, schedule)
                            except Exception:
                                pass
                            # Force-insert united lessons into each sub-group NOW,
                            # before their autofill runs, so preserve_existing keeps them.
                            g_obj = groups_map.get(name)
                            if g_obj and g_obj.get('sub_groups'):
                                for sub_name in g_obj['sub_groups']:
                                    try:
                                        excel_service.force_merge_into_group_schedule(sub_name, schedule)
                                    except Exception:
                                        pass
                        if success:
                            break
                        attempt += 1
                    g_obj = groups_map.get(name)
                    rebuild_log.append({'group': name, 'success': bool(success), 'attempts': attempt, 'info': last_info, 'is_united': True})
                except Exception as e:
                    rebuild_log.append({'group': name, 'success': False, 'errors': [str(e)], 'is_united': True})

            # ── Phase 3: build regular groups (preserve united slots) ─────────────
            # preserve_existing=True means autofill keeps the force-inserted united
            # lessons and only fills the remaining empty slots with the group's own
            # subjects.  Teacher busy_map also sees the saved united schedules,
            # so there are no teacher double-bookings.
            for name in regular_names_ordered:
                if not name:
                    continue
                try:
                    # log group start
                    try:
                        with open(os.path.join('uploads', 'rebuild_all.log'), 'a', encoding='utf-8') as lf:
                            lf.write(f"{datetime.now().isoformat()} | group_start | {name}\n")
                    except Exception:
                        pass
                    outer_max = int(config.get('max_autofill_retries', max_autofill_retries)) if config else max_autofill_retries
                    attempt = 0
                    success = False
                    last_info = None
                    while attempt < outer_max and not success:
                        try:
                            if excel_service:
                                excel_service.load()
                        except Exception:
                            pass
                        # preserve_existing=True: keep force-inserted united lessons,
                        # fill remaining empty slots with this group's own subjects.
                        try:
                            success, schedule, info = autofill_service.autofill_group(name, max_retries=1, preserve_existing=True, subject_order=subject_order)
                        except TypeError:
                            success, schedule, info = autofill_service.autofill_group(name)

                        last_info = info

                        if schedule and any(schedule.values()):
                            try:
                                excel_service.save_group_schedule(name, schedule)
                            except Exception:
                                pass

                        if success:
                            break

                        # If incomplete info provided, try to change priorities: move failing subjects to top
                        try:
                            incomplete = []
                            if isinstance(info, dict):
                                incomplete = info.get('incomplete', []) or []
                            if incomplete:
                                subs = excel_service.get_subjects() or []
                                fail_names = set([it.get('subject') for it in incomplete if isinstance(it, dict) and it.get('subject')])
                                if fail_names:
                                    front = [s for s in subs if s.get('name') in fail_names]
                                    rest = [s for s in subs if s.get('name') not in fail_names]
                                    excel_service.save_subjects(front + rest)
                        except Exception:
                            pass

                        attempt += 1

                    # log group end
                    try:
                        with open(os.path.join('uploads', 'rebuild_all.log'), 'a', encoding='utf-8') as lf:
                            lf.write(f"{datetime.now().isoformat()} | group_end | {name} | attempts={attempt} | success={bool(success)}\n")
                    except Exception:
                        pass

                    rebuild_log.append({'group': name, 'success': bool(success), 'attempts': attempt, 'info': last_info, 'is_united': False})
                except Exception as e:
                    rebuild_log.append({'group': name, 'success': False, 'errors': [str(e)], 'is_united': False})
                    continue

            # ensure teacher schedules rebuilt
            try:
                if excel_service:
                    excel_service.rebuild_teacher_schedules()
            except Exception:
                pass

            return jsonify({'success': True, 'groups': len(groups), 'log': rebuild_log})
        else:
            return jsonify({'success': False, 'error': 'autofill service not available'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/admin/export', methods=['GET'])
def export_excel():
    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Configuration
    cfg = excel_service.get_config() if excel_service else {}
    ws = wb.create_sheet('Configuration')
    ws.append(['Setting', 'Value'])
    for k, v in (cfg.items() if cfg else []):
        ws.append([k, v])

    # Language Texts
    texts = excel_service.get_language_texts() if excel_service else {}
    ws = wb.create_sheet('Language Texts')
    ws.append(['Language', 'app_name', 'autofill_direction', 'weekdays'])
    for lang, vals in (texts.items() if texts else []):
        ws.append([lang, vals.get('app_name', ''), vals.get('autofill_direction', ''), ','.join(vals.get('weekdays', []))])

    # Weekdays
    ws = wb.create_sheet('Weekdays')
    ws.append(['Order', 'Weekday', 'Short'])
    weekdays_str = cfg.get('WEEKDAYS', '') if cfg else ''
    if weekdays_str:
        days = [d.strip() for d in weekdays_str.split(',') if d.strip()]
        for idx, d in enumerate(days, start=1):
            ws.append([idx, d, d[:2]])

    # Time Slots
    time_slots = excel_service.get_time_slots() if excel_service else {}
    ws = wb.create_sheet('Time Slots')
    ws.append(['Lesson', 'Time Range'])
    for lesson in sorted(time_slots.keys()):
        ws.append([lesson, time_slots[lesson]])

    # Teachers (match ExcelExamples header exactly)
    ws = wb.create_sheet('Teachers')
    ws.append([
        'Teacher Name',
        'Subjects (Name:Hours:Group)',
        'Total Hours',
        'Availability (Day:Start-End)',
        'Teachers Weekly Hours (Day:Start-End)'
    ])
    teachers = excel_service.get_teachers() if excel_service else []
    for t in teachers:
        subjects = t.get('subjects', [])
        check_in = t.get('check_in_hours', {})
        check_out = t.get('check_out_hours', {})

        # weekly hours string (Day:Start-End; ...)
        weekly_parts = []
        for d in (check_in or {}):
            for i, v in enumerate(check_in.get(d, [])):
                out = check_out.get(d, [])
                outv = out[i] if i < len(out) else ''
                weekly_parts.append(f"{d}:{v}-{outv}")
        weekly = '; '.join(weekly_parts)

        # availability string (Day:min-max; ...)
        avail_str = ''
        try:
            avail = t.get('available_slots', {}) or {}
            avail_parts = [f"{d}:{min(v)}-{max(v)}" for d, v in avail.items() if v]
            avail_str = '; '.join(avail_parts)
        except Exception:
            avail_str = ''

        # subjects string "Name:Hours:Group"
        subj_parts = []
        total = 0
        for s in (subjects or []):
            if isinstance(s, dict):
                name = s.get('name', '')
                hours = int(s.get('hours', 0) or 0)
                group = s.get('group', '')
            else:
                name = s or ''
                hours = 0
                group = ''
            subj_parts.append(f"{name}:{hours}:{group}")
            try:
                total += int(hours)
            except Exception:
                pass
        subjects_str = '; '.join(subj_parts)

        ws.append([t.get('name', ''), subjects_str, total, avail_str, weekly])

    # Groups (match ExcelExamples header + IsUnited/SubGroups for united-group support)
    ws = wb.create_sheet('Groups')
    ws.append(['Group Name', 'Comment', 'IsUnited', 'SubGroups'])
    groups = excel_service.get_groups() if excel_service else []
    for g in groups:
        subjects_str = '; '.join(g.get('subjects', [])) if isinstance(g, dict) else ''
        name = g.get('name', '') if isinstance(g, dict) else (g or '')
        is_united = 1 if (isinstance(g, dict) and g.get('is_united')) else 0
        sub_groups_str = '; '.join(g.get('sub_groups', [])) if isinstance(g, dict) else ''
        ws.append([name, subjects_str, is_united, sub_groups_str])

    # Subjects (match ExcelExamples header)
    ws = wb.create_sheet('Subjects')
    ws.append(['Subject Name', 'Hours per Week', 'Group'])
    subjects = excel_service.get_subjects() if excel_service else []
    for s in subjects:
        ws.append([
            s.get('name', ''),
            s.get('hours_per_week', 0),
            s.get('group', '')
        ])

    # Group Schedules
    ws = wb.create_sheet('Group Schedules')
    ws.append(['Group', 'Day', 'Lesson', 'Subject', 'Teacher', 'Color BG', 'Color FG'])
    group_schedules = excel_service.get_group_schedules() if excel_service else {}
    for group_name, days in (group_schedules.items() if group_schedules else []):
        for day, lessons in (days.items() if days else []):
            for lesson_num, ld in (lessons.items() if lessons else []):
                ws.append([
                    group_name,
                    day,
                    lesson_num,
                    ld.get('subject', '') if isinstance(ld, dict) else '',
                    ld.get('teacher', '') if isinstance(ld, dict) else '',
                    ld.get('color_bg', '#FFFFFF') if isinstance(ld, dict) else '#FFFFFF',
                    ld.get('color_fg', '#000000') if isinstance(ld, dict) else '#000000'
                ])

    # Teacher Schedules
    ws = wb.create_sheet('Teacher Schedules')
    ws.append(['Teacher', 'Day', 'Lesson', 'Subject', 'Group', 'Color BG', 'Color FG'])
    teacher_schedules = excel_service.get_teacher_schedules() if excel_service else {}
    for teacher_name, days in (teacher_schedules.items() if teacher_schedules else []):
        for day, lessons in (days.items() if days else []):
            for lesson_num, ld in (lessons.items() if lessons else []):
                ws.append([
                    teacher_name,
                    day,
                    lesson_num,
                    ld.get('subject', '') if isinstance(ld, dict) else '',
                    ld.get('group', '') if isinstance(ld, dict) else '',
                    ld.get('color_bg', '#FFFFFF') if isinstance(ld, dict) else '#FFFFFF',
                    ld.get('color_fg', '#000000') if isinstance(ld, dict) else '#000000'
                ])

    # Generate matrix-style Group_<Name> and Teacher_<Name> sheets with colored lesson cells
    try:
        from openpyxl.styles import PatternFill, Font

        # Helper to convert '#RRGGBB' -> 'FFRRGGBB' (ARGB) for openpyxl
        def _to_argb(hex_color):
            if not hex_color:
                return 'FFFFFFFF'
            c = str(hex_color).strip()
            if c.startswith('#'):
                c = c[1:]
            if len(c) == 6:
                return 'FF' + c.upper()
            # fallback
            return 'FFFFFFFF'

        # Determine weekdays and lessons order
        cfg_weekdays = cfg.get('WEEKDAYS', '') if cfg else ''
        weekdays = [d.strip() for d in cfg_weekdays.split(',')] if cfg_weekdays else []
        time_slots = excel_service.get_time_slots() if excel_service else {}
        lesson_nums = sorted(time_slots.keys()) if time_slots else []

        # Build Group_ sheets (transposed: rows = lessons, columns = weekdays)
        groups = excel_service.get_groups() if excel_service else []
        grp_names = [g.get('name') if isinstance(g, dict) else (g or '') for g in groups]
        for gname in grp_names:
            sname = f'Group_{gname}'
            if not gname or sname in wb.sheetnames:
                continue
            tgt = wb.create_sheet(sname)
            # header: first cell = 'Lesson/Day', then weekdays as columns
            header = ['Lesson/Day']
            if weekdays:
                for wd in weekdays:
                    header.append(wd)
            else:
                # fallback to columns from group schedule
                gs = group_schedules.get(gname, {}) if group_schedules else {}
                wd_list = sorted(list(gs.keys())) if gs else []
                for wd in wd_list:
                    header.append(wd)
            tgt.append(header)

            # fill rows per lesson number
            gs = group_schedules.get(gname, {}) if group_schedules else {}
            for ln in lesson_nums:
                row = [time_slots.get(ln, str(ln))]
                for wd in (weekdays or list(gs.keys())):
                    cell_val = ''
                    if gs and wd in gs and ln in gs[wd]:
                        ld = gs[wd][ln]
                        cell_val = ld.get('subject', '') if isinstance(ld, dict) else ''
                    row.append(cell_val)
                tgt.append(row)
                # apply styles for this lesson-row
                r_idx = tgt.max_row
                for c_idx, wd in enumerate((weekdays or list(gs.keys())), start=2):
                    cell = tgt.cell(row=r_idx, column=c_idx)
                    if gs and wd in gs and ln in gs[wd]:
                        ld = gs[wd][ln]
                        bg = _to_argb(ld.get('color_bg', '#FFFFFF') if isinstance(ld, dict) else '#FFFFFF')
                        fg = _to_argb(ld.get('color_fg', '#000000') if isinstance(ld, dict) else '#000000')
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                        cell.font = Font(color=fg)

        # Build Teacher_ sheets (transposed: rows = lessons, columns = weekdays)
        # First, create a map of teacher availability
        teacher_availability = {}
        for t in teachers:
            tname = t.get('name', '')
            if tname:
                teacher_availability[tname] = t.get('available_slots', {})
        
        for tname, tsched in (teacher_schedules.items() if teacher_schedules else []):
            sname = f'Teacher_{tname}'
            if not tname or sname in wb.sheetnames:
                continue
            tgt = wb.create_sheet(sname)
            header = ['Lesson/Day']
            if weekdays:
                for wd in weekdays:
                    header.append(wd)
            else:
                wd_list = sorted(list(tsched.keys())) if tsched else []
                for wd in wd_list:
                    header.append(wd)
            tgt.append(header)

            # Get this teacher's available slots
            avail_slots = teacher_availability.get(tname, {})

            for ln in lesson_nums:
                row = [time_slots.get(ln, str(ln))]
                for wd in (weekdays or list(tsched.keys())):
                    cell_val = ''
                    if tsched and wd in tsched and ln in tsched[wd]:
                        ld = tsched[wd][ln]
                        cell_val = ld.get('subject', '') if isinstance(ld, dict) else ''
                    else:
                        # Check if this slot is in teacher's available hours
                        if avail_slots and wd in avail_slots and ln in avail_slots[wd]:
                            cell_val = 'Empty'
                        else:
                            cell_val = ''
                    row.append(cell_val)
                tgt.append(row)
                # style cells for this lesson-row
                r_idx = tgt.max_row
                for c_idx, wd in enumerate((weekdays or list(tsched.keys())), start=2):
                    cell = tgt.cell(row=r_idx, column=c_idx)
                    if tsched and wd in tsched and ln in tsched[wd]:
                        ld = tsched[wd][ln]
                        bg = _to_argb(ld.get('color_bg', '#FFFFFF') if isinstance(ld, dict) else '#FFFFFF')
                        fg = _to_argb(ld.get('color_fg', '#000000') if isinstance(ld, dict) else '#000000')
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                        cell.font = Font(color=fg)
    except Exception:
        # non-critical: coloring/generation optional
        pass

    # Copy any additional Group_/Teacher_ sheets present in the application's Excel workbook,
    # but only if they contain non-empty cells (beyond header). If not present there,
    # look through all files in ExcelExamples and copy non-empty sheets.
    try:
        existing = set(wb.sheetnames)

        def sheet_has_content(ws):
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                return False
            for r in rows[1:]:
                # skip first column (lesson label)
                for c in r[1:]:
                    if c is not None and str(c).strip() != '':
                        return True
            return False

        # 1) Prefer sheets from the running application's workbook
        if excel_service and getattr(excel_service, 'wb', None):
            src_wb = excel_service.wb
            for sname in src_wb.sheetnames:
                if sname in existing:
                    continue
                if not (sname.startswith('Group_') or sname.startswith('Teacher_')):
                    continue
                try:
                    src_ws = src_wb[sname]
                    if not sheet_has_content(src_ws):
                        continue
                    tgt_ws = wb.create_sheet(sname)
                    for row in src_ws.iter_rows(values_only=True):
                        tgt_ws.append(list(row))
                    existing.add(sname)
                except Exception:
                    continue

        # Fallback logic removed - only export actual database data, not examples
    except Exception:
        # non-critical: if this fails we still return core sheets
        pass

    # Save file
    os.makedirs('uploads', exist_ok=True)
    fname = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join('uploads', fname)
    wb.save(path)
    return jsonify({'success': True, 'filename': fname})


@admin_bp.route('/api/admin/priorities-data', methods=['GET'])
def priorities_data():
    try:
        all_groups = excel_service.get_groups() if excel_service else []
        # United groups are virtual — exclude from rebuild priorities
        groups = [g for g in all_groups if not g.get('is_united')]
        subjects = excel_service.get_subjects() if excel_service else []
        teachers = excel_service.get_teachers() if excel_service else []
        cfg = excel_service.get_config() if excel_service else {}
        weekdays_str = cfg.get('WEEKDAYS', '') if cfg else ''
        weekdays = [d.strip() for d in weekdays_str.split(',')] if weekdays_str else []

        # Auto-sort subjects by complexity: most constrained first.
        # Priority tiers (lower sort key = earlier):
        #   Tier 0: subjects of united groups (must place first — shared across sub-groups)
        #   Tier 1: multi-teacher subjects (subgroups feature — very constrained)
        #   Tier 2: subjects with a specific group assigned, sorted by hours desc + teacher scarcity
        #   Tier 3: subjects with no group (group='') — these are generic/orphan, place last

        # Build set of united group names for fast lookup
        united_group_names = set(g.get('name', '') for g in all_groups if isinstance(g, dict) and g.get('is_united'))

        # Build (subject_name, group_name) -> [teacher_objects] map
        teacher_subj_map = {}
        for t in teachers:
            for ts in t.get('subjects', []):
                key = (ts.get('name', ''), ts.get('group', ''))
                teacher_subj_map.setdefault(key, []).append(t)

        def _complexity(s):
            s_name = s.get('name', '') if isinstance(s, dict) else ''
            s_group = s.get('group', '') if isinstance(s, dict) else ''
            hours = int(s.get('hours_per_week', 0) or 0) if isinstance(s, dict) else 0

            # Subjects with no group go to the very bottom
            if not s_group:
                return (3, 0, 0)

            matching = teacher_subj_map.get((s_name, s_group), []) + teacher_subj_map.get((s_name, ''), [])
            # deduplicate by teacher name
            seen = set()
            unique_teachers = []
            for t in matching:
                tn = t.get('name', '')
                if tn not in seen:
                    seen.add(tn)
                    unique_teachers.append(t)
            teacher_count = len(unique_teachers)
            total_slots = sum(len(v) for t in unique_teachers for v in t.get('available_slots', {}).values() if v)

            is_united_subj = s_group in united_group_names
            is_multi_teacher = teacher_count > 1

            if is_united_subj:
                tier = 0
            elif is_multi_teacher:
                tier = 1
            else:
                tier = 2

            # Within tier: more hours first, then fewer slots (harder = earlier)
            # negate hours so higher hours sorts first; negate total_slots so fewer slots sorts first
            return (tier, -hours, total_slots if total_slots > 0 else 99999)

        try:
            subjects_sorted = sorted(subjects, key=_complexity)
        except Exception:
            subjects_sorted = subjects

        return jsonify({'success': True, 'groups': groups, 'subjects': subjects_sorted, 'teachers': teachers, 'weekdays': weekdays})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/admin/debug-data', methods=['GET'])
def debug_data():
    """Return debug information about the active Excel file and data counts."""
    try:
        file_path = None
        if excel_service:
            file_path = getattr(excel_service, 'file_path', None)
        teachers = excel_service.get_teachers() if excel_service else []
        groups = excel_service.get_groups() if excel_service else []
        subjects = excel_service.get_subjects() if excel_service else []
        group_schedules = excel_service.get_group_schedules() if excel_service else {}
        teacher_schedules = excel_service.get_teacher_schedules() if excel_service else {}

        return jsonify({
            'success': True,
            'file_path': file_path,
            'counts': {
                'teachers': len(teachers),
                'groups': len(groups),
                'subjects': len(subjects),
                'group_schedules_groups': len(group_schedules.keys()),
                'teacher_schedules_teachers': len(teacher_schedules.keys())
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/api/color/<path:subject_name>', methods=['GET'])
def get_color(subject_name):
    """Return background/foreground color for a subject using ColorService."""
    try:
        cs = None
        if autofill_service and getattr(autofill_service, 'color_service', None):
            cs = autofill_service.color_service
        else:
            from services.color_service import ColorService
            cs = ColorService()
        bg, fg = cs.get_color(subject_name)
        return jsonify({'bg': bg, 'fg': fg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
