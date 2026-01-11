import os
import openpyxl
from openpyxl import Workbook, load_workbook
from models import Teacher, Group, Subject


class ExcelService:
    def __init__(self, file_path=None, username=None):
        """
        Initialize ExcelService
        If username is provided, use user-specific file: data/user_{username}.xlsx
        Otherwise use file_path
        """
        if username:
            data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
            os.makedirs(data_dir, exist_ok=True)
            self.file_path = os.path.join(data_dir, f'user_{username}.xlsx')
        else:
            self.file_path = file_path
        
        self.wb = None
        self.load()
    
    def load(self):
        """Load Excel file or create if not exists"""
        if os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self._create_default()
    
    def _create_default(self):
        """Create default Excel file with 9 sheets (including Language Texts)"""
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Sheet 1: Configuration
        ws_config = self.wb.create_sheet('Configuration')
        ws_config.append(['Setting', 'Value'])
        ws_config.append(['GUI_LANGUAGE', 'English'])
        ws_config.append(['LANGUAGES_LIST', 'English,Hebrew,Russian'])
        ws_config.append(['app_name', 'School Schedule Editor'])
        ws_config.append(['app_size', '1500x900'])
        ws_config.append(['autofill_direction', 'Left_to_Right'])
        ws_config.append(['max_autofill_retries', '2'])
        ws_config.append(['max_sequence_lessons', '2'])
        ws_config.append(['max_per_day', '3'])
        ws_config.append(['lessons', '10'])
        ws_config.append(['unavailable_slot', 'XXXXXXXX'])
        ws_config.append(['WEEKDAYS', 'Sunday,Monday,Tuesday,Wednesday,Thursday,Friday,Saturday'])
        
        # Sheet 2: Language Texts
        ws_texts = self.wb.create_sheet('Language Texts')
        ws_texts.append(['Language', 'app_name', 'autofill_direction', 'weekdays'])
        ws_texts.append([
            'English', 
            'School Schedule Editor', 
            'Left_to_Right',
            'Sunday,Monday,Tuesday,Wednesday,Thursday,Friday,Saturday'
        ])
        ws_texts.append([
            'Hebrew', 
            'עורך מערכת שעות', 
            'Right_to_Left',
            'יום שישי,יום חמישי,יום רביעי,יום שלישי,יום שני,יום ראשון'
        ])
        ws_texts.append([
            'Russian', 
            'Редактор школьного расписания', 
            'Left_to_Right',
            'Понедельник,Вторник,Среда,Четверг,Пятница,Суббота'
        ])
        
        # Sheet 3: Weekdays
        ws_weekdays = self.wb.create_sheet('Weekdays')
        ws_weekdays.append(['Order', 'Weekday', 'Short'])
        weekdays = [
            (1, 'Sunday', 'Su'),
            (2, 'Monday', 'M'),
            (3, 'Tuesday', 'T'),
            (4, 'Wednesday', 'W'),
            (5, 'Thursday', 'Th'),
            (6, 'Friday', 'F'),
            (7, 'Saturday', 'Sa')
        ]
        for wd in weekdays:
            ws_weekdays.append(wd)
        
        # Sheet 4: Time Slots
        ws_time = self.wb.create_sheet('Time Slots')
        ws_time.append(['Lesson', 'Time Range'])
        time_slots = {
            1: '09:15-10:00',
            2: '10:00-10:40',
            3: '10:45-11:30',
            4: '11:30-12:15',
            5: '12:20-13:00',
            6: '14:00-14:45',
            7: '14:55-15:40',
            8: '15:50-16:35',
            9: '16:40-17:25',
            10: '17:30-18:15'
        }
        for lesson, time_range in time_slots.items():
            ws_time.append([lesson, time_range])
        
        # Sheet 5: Teachers
        ws_teachers = self.wb.create_sheet('Teachers')
        ws_teachers.append(['Name', 'Subject', 'Hours', 'Group', 'Teachers Weekly Hours', 'Availability'])
        
        # Sheet 6: Groups
        ws_groups = self.wb.create_sheet('Groups')
        ws_groups.append(['Name', 'Comments'])
        
        # Sheet 7: Subjects
        ws_subjects = self.wb.create_sheet('Subjects')
        ws_subjects.append(['Name', 'Group', 'Hours Per Week', 'Teacher'])
        
        # Sheet 8: Group Schedules
        ws_group_sched = self.wb.create_sheet('Group Schedules')
        ws_group_sched.append(['Group', 'Day', 'Lesson', 'Subject', 'Teacher', 'Color BG', 'Color FG'])
        
        # Sheet 9: Teacher Schedules
        ws_teacher_sched = self.wb.create_sheet('Teacher Schedules')
        ws_teacher_sched.append(['Teacher', 'Day', 'Lesson', 'Subject', 'Group', 'Color BG', 'Color FG'])
        
        # Ensure data directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        self.wb.save(self.file_path)
    
    def get_config(self):
        """Read Configuration sheet"""
        if 'Configuration' not in self.wb.sheetnames:
            return {}
        
        sheet = self.wb['Configuration']
        config = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config[row[0]] = row[1]
        return config
    
    def save_config(self, config):
        """Save configuration"""
        if 'Configuration' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Configuration')
            ws.append(['Setting', 'Value'])
        else:
            ws = self.wb['Configuration']
            # Clear existing data
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
        
        # Write new config
        row_idx = 2
        for key, value in config.items():
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)
            row_idx += 1
        
        self.save()
    
    def get_language_texts(self):
        """
        Read Language Texts sheet
        Returns: dict with structure like source_consts.py TEXTS
        {
            'English': {
                'app_name': 'School Schedule Editor',
                'autofill_direction': 'Left_to_Right',
                'weekdays': ['Sunday', 'Monday', ...]
            },
            ...
        }
        """
        if 'Language Texts' not in self.wb.sheetnames:
            return {}
        
        sheet = self.wb['Language Texts']
        texts = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            language = row[0]
            app_name = row[1] if len(row) > 1 else ''
            autofill_direction = row[2] if len(row) > 2 else 'Left_to_Right'
            weekdays_str = row[3] if len(row) > 3 else ''
            
            texts[language] = {
                'app_name': app_name,
                'autofill_direction': autofill_direction,
                'weekdays': weekdays_str.split(',') if weekdays_str else []
            }
        
        return texts
    
    def get_time_slots(self):
        """Get time slots from Time Slots sheet"""
        if 'Time Slots' not in self.wb.sheetnames:
            return {}
        
        sheet = self.wb['Time Slots']
        time_slots = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                time_slots[int(row[0])] = row[1]
        return time_slots
    
    def save_time_slots(self, time_slots):
        """
        Save time slots to Time Slots sheet
        Args:
            time_slots: dict {lesson_num: time_range}, e.g., {1: '09:15-10:00', 2: '10:00-10:40'}
        """
        if 'Time Slots' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Time Slots')
            ws.append(['Lesson', 'Time Range'])
        else:
            ws = self.wb['Time Slots']
            # Clear existing data (keep header)
            ws.delete_rows(2, ws.max_row)
        
        # Write time slots sorted by lesson number
        for lesson_num in sorted(time_slots.keys()):
            ws.append([lesson_num, time_slots[lesson_num]])
        
        self.save()
    
    def get_teachers(self):
        """Read Teachers sheet with PRIORITY logic"""
        if 'Teachers' not in self.wb.sheetnames:
            return []
        
        sheet = self.wb['Teachers']
        teachers = {}
        time_slots = self.get_time_slots()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            name = row[0]
            subject = row[1] if len(row) > 1 else None
            hours = row[2] if len(row) > 2 else 0
            group = row[3] if len(row) > 3 else None
            weekly_hours = row[4] if len(row) > 4 else None
            availability = row[5] if len(row) > 5 else None
            
            if name not in teachers:
                teachers[name] = {
                    'name': name,
                    'subjects': [],
                    'available_slots': {},
                    'check_in_hours': {},
                    'check_out_hours': {}
                }
            
            if subject:
                teachers[name]['subjects'].append({
                    'name': subject,
                    'hours': int(hours) if hours else 0,
                    'group': group
                })
            
            # PRIORITY: Teachers Weekly Hours > Availability
            if weekly_hours:
                check_in, check_out = self._parse_weekly_hours(weekly_hours)
                teachers[name]['check_in_hours'] = check_in
                teachers[name]['check_out_hours'] = check_out
                # Recalculate available_slots from times
                teachers[name]['available_slots'] = self._convert_times_to_lessons(
                    check_in, check_out, time_slots
                )
            elif availability:
                teachers[name]['available_slots'] = self._parse_availability(availability)
        
        return list(teachers.values())
    
    def _parse_weekly_hours(self, weekly_hours_str):
        """
        Parse: "Monday:08:00-12:00; Monday:13:00-17:00; Tuesday:09:00-15:00"
        Returns: ({Monday: ['08:00', '13:00'], Tuesday: ['09:00']}, 
                  {Monday: ['12:00', '17:00'], Tuesday: ['15:00']})
        """
        check_in = {}
        check_out = {}
        
        if not weekly_hours_str:
            return check_in, check_out
        
        for entry in str(weekly_hours_str).split(';'):
            entry = entry.strip()
            if not entry or ':' not in entry:
                continue
            
            parts = entry.split(':')
            if len(parts) < 2:
                continue
            
            day = parts[0].strip()
            time_part = ':'.join(parts[1:])
            
            if '-' not in time_part:
                continue
            
            time_range = time_part.strip()
            if '-' in time_range:
                start, end = time_range.split('-')
                start = start.strip()
                end = end.strip()
                
                if day not in check_in:
                    check_in[day] = []
                    check_out[day] = []
                
                check_in[day].append(start)
                check_out[day].append(end)
        
        return check_in, check_out
    
    def _convert_times_to_lessons(self, check_in, check_out, time_slots):
        """Convert HH:MM ranges to lesson numbers using TIME_SLOTS"""
        available_slots = {}
        for day in check_in:
            available_slots[day] = []

            for start_time, end_time in zip(check_in[day], check_out[day]):
                try:
                    # parse check in/out to minutes
                    ch_h, ch_m = map(int, start_time.split(':'))
                    co_h, co_m = map(int, end_time.split(':'))
                    checkin_minutes = ch_h * 60 + ch_m
                    checkout_minutes = co_h * 60 + co_m
                except Exception:
                    # invalid format -> skip this interval
                    continue

                # iterate lessons and include only those lessons fully contained
                for lesson_num, time_range in time_slots.items():
                    if '-' not in time_range:
                        continue
                    slot_start_str, slot_end_str = [s.strip() for s in time_range.split('-', 1)]
                    try:
                        ss_h, ss_m = map(int, slot_start_str.split(':'))
                        se_h, se_m = map(int, slot_end_str.split(':'))
                        slot_start_minutes = ss_h * 60 + ss_m
                        slot_end_minutes = se_h * 60 + se_m
                    except Exception:
                        continue

                    # include lesson only if lesson fully inside [checkin, checkout]
                    if checkin_minutes <= slot_start_minutes and checkout_minutes >= slot_end_minutes:
                        if lesson_num not in available_slots[day]:
                            available_slots[day].append(lesson_num)

        return available_slots
    
    def _parse_availability(self, availability_str):
        """
        Parse: "Monday:1-5; Tuesday:1-3"
        Returns: {Monday: [1,2,3,4,5], Tuesday: [1,2,3]}
        """
        available_slots = {}
        
        if not availability_str:
            return available_slots
        
        for entry in str(availability_str).split(';'):
            entry = entry.strip()
            if not entry or ':' not in entry:
                continue
            
            day, lessons = entry.split(':', 1)
            day = day.strip()
            lessons = lessons.strip()
            
            if '-' in lessons:
                start, end = lessons.split('-')
                available_slots[day] = list(range(int(start), int(end) + 1))
            else:
                available_slots[day] = [int(lessons)]
        
        return available_slots
    
    def save_teachers(self, teachers):
        """Save teachers to Excel"""
        if 'Teachers' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Teachers')
            ws.append(['Name', 'Subject', 'Hours', 'Group', 'Teachers Weekly Hours', 'Availability'])
        else:
            ws = self.wb['Teachers']
            # Clear existing data
            ws.delete_rows(2, ws.max_row)
        
        # Write teachers
        for teacher in teachers:
            name = teacher['name']
            subjects = teacher.get('subjects', [])
            check_in = teacher.get('check_in_hours', {})
            check_out = teacher.get('check_out_hours', {})
            available_slots = teacher.get('available_slots', {})
            
            # Convert check_in/check_out to weekly hours string
            weekly_hours = self._format_weekly_hours(check_in, check_out)
            
            # Convert available_slots to availability string
            availability = self._format_availability(available_slots)
            
            if subjects:
                for subject in subjects:
                    ws.append([
                        name,
                        subject.get('name', ''),
                        subject.get('hours', 0),
                        subject.get('group', ''),
                        weekly_hours,
                        availability
                    ])
            else:
                ws.append([name, '', 0, '', weekly_hours, availability])
        
        self.save()
    
    def _format_weekly_hours(self, check_in, check_out):
        """Format check_in/check_out to string"""
        if not check_in:
            return ''
        
        parts = []
        for day in check_in:
            for start, end in zip(check_in[day], check_out.get(day, [])):
                parts.append(f'{day}:{start}-{end}')
        
        return '; '.join(parts)
    
    def _format_availability(self, available_slots):
        """Format available_slots to string"""
        if not available_slots:
            return ''
        
        parts = []
        for day, lessons in available_slots.items():
            if lessons:
                min_lesson = min(lessons)
                max_lesson = max(lessons)
                parts.append(f'{day}:{min_lesson}-{max_lesson}')
        
        return '; '.join(parts)
    
    def get_groups(self):
        """Read Groups sheet"""
        if 'Groups' not in self.wb.sheetnames:
            return []
        
        sheet = self.wb['Groups']
        groups = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                name = row[0]
                subjects = row[1].split(';') if len(row) > 1 and row[1] else []
                groups.append({
                    'name': name,
                    'subjects': [s.strip() for s in subjects if s.strip()]
                })
        return groups
    
    def save_groups(self, groups):
        """Save groups to Excel"""
        if 'Groups' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Groups')
            ws.append(['Name', 'Comments'])
        else:
            ws = self.wb['Groups']
            ws.delete_rows(2, ws.max_row)
        
        for group in groups:
            subjects_str = '; '.join(group.get('subjects', []))
            ws.append([group['name'], subjects_str])
        
        self.save()
    
    def get_subjects(self):
        """Read Subjects sheet"""
        if 'Subjects' not in self.wb.sheetnames:
            return []
        
        sheet = self.wb['Subjects']
        subjects = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                subjects.append({
                    'name': row[0],
                    'group': row[1] if len(row) > 1 else '',
                    'hours_per_week': int(row[2]) if len(row) > 2 and row[2] else 0,
                    'teacher': row[3] if len(row) > 3 else ''
                })
        return subjects
    
    def save_subjects(self, subjects):
        """Save subjects to Excel"""
        if 'Subjects' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Subjects')
            ws.append(['Name', 'Group', 'Hours Per Week', 'Teacher'])
        else:
            ws = self.wb['Subjects']
            ws.delete_rows(2, ws.max_row)
        
        for subject in subjects:
            ws.append([
                subject['name'],
                subject.get('group', ''),
                subject.get('hours_per_week', 0),
                subject.get('teacher', '')
            ])
        
        self.save()

    def update_subject_in_teachers(self, old_name, new_subject):
        """Update Teachers sheet rows that reference a subject name.
        Args:
            old_name: original subject name to find
            new_subject: dict with keys 'name', 'group', 'hours_per_week'
        """
        if 'Teachers' not in self.wb.sheetnames:
            return

        ws = self.wb['Teachers']
        for row in ws.iter_rows(min_row=2):
            cell_subject = row[1].value if len(row) > 1 else None
            if cell_subject and cell_subject == old_name:
                # Update subject name, hours and group columns
                row[1].value = new_subject.get('name', old_name)
                if len(row) > 2:
                    row[2].value = new_subject.get('hours_per_week', row[2].value)
                if len(row) > 3:
                    row[3].value = new_subject.get('group', row[3].value)

        self.save()

    def remove_subject_from_teachers(self, name):
        """Remove subject references from Teachers sheet (set subject empty and hours 0).
        Args:
            name: subject name to remove
        """
        if 'Teachers' not in self.wb.sheetnames:
            return

        ws = self.wb['Teachers']
        for row in ws.iter_rows(min_row=2):
            cell_subject = row[1].value if len(row) > 1 else None
            if cell_subject and cell_subject == name:
                row[1].value = ''
                if len(row) > 2:
                    row[2].value = 0
                if len(row) > 3:
                    row[3].value = ''

        self.save()
    
    def get_group_schedules(self):
        """Read Group Schedules sheet"""
        if 'Group Schedules' not in self.wb.sheetnames:
            return {}
        
        sheet = self.wb['Group Schedules']
        schedules = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            group = row[0]
            day = row[1]
            lesson = int(row[2]) if row[2] else 0
            subject = row[3] if len(row) > 3 else ''
            teacher = row[4] if len(row) > 4 else ''
            color_bg = row[5] if len(row) > 5 else '#FFFFFF'
            color_fg = row[6] if len(row) > 6 else '#000000'
            
            if group not in schedules:
                schedules[group] = {}
            if day not in schedules[group]:
                schedules[group][day] = {}
            
            schedules[group][day][lesson] = {
                'subject': subject,
                'teacher': teacher,
                'group': group,
                'color_bg': color_bg,
                'color_fg': color_fg
            }
        
        return schedules
    
    def save_group_schedule(self, group_name, schedule):
        """Save a group's schedule"""
        if 'Group Schedules' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Group Schedules')
            ws.append(['Group', 'Day', 'Lesson', 'Subject', 'Teacher', 'Color BG', 'Color FG'])
        else:
            ws = self.wb['Group Schedules']
        
        # Remove existing schedule for this group
        rows_to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == group_name:
                rows_to_delete.append(idx)
        
        for idx in reversed(rows_to_delete):
            ws.delete_rows(idx)
        
        # Add new schedule
        for day, lessons in schedule.items():
            for lesson_num, lesson_data in lessons.items():
                ws.append([
                    group_name,
                    day,
                    lesson_num,
                    lesson_data.get('subject', ''),
                    lesson_data.get('teacher', ''),
                    lesson_data.get('color_bg', '#FFFFFF'),
                    lesson_data.get('color_fg', '#000000')
                ])
        
        self.save()
    
    def get_teacher_schedules(self):
        """Read Teacher Schedules sheet"""
        if 'Teacher Schedules' not in self.wb.sheetnames:
            return {}
        
        sheet = self.wb['Teacher Schedules']
        schedules = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            teacher = row[0]
            day = row[1]
            lesson = int(row[2]) if row[2] else 0
            subject = row[3] if len(row) > 3 else ''
            group = row[4] if len(row) > 4 else ''
            color_bg = row[5] if len(row) > 5 else '#FFFFFF'
            color_fg = row[6] if len(row) > 6 else '#000000'
            
            if teacher not in schedules:
                schedules[teacher] = {}
            if day not in schedules[teacher]:
                schedules[teacher][day] = {}
            
            schedules[teacher][day][lesson] = {
                'subject': subject,
                'teacher': teacher,
                'group': group,
                'color_bg': color_bg,
                'color_fg': color_fg
            }
        
        return schedules
    
    def rebuild_teacher_schedules(self):
        """Rebuild teacher schedules from group schedules"""
        group_schedules = self.get_group_schedules()
        teacher_schedules = {}
        
        for group_name, group_schedule in group_schedules.items():
            for day, lessons in group_schedule.items():
                for lesson_num, lesson_data in lessons.items():
                    teacher_field = lesson_data.get('teacher', '') or ''
                    if not teacher_field:
                        continue
                    # teacher_field may contain multiple teacher names separated by ';' or ','
                    if isinstance(teacher_field, str):
                        teachers = [t.strip() for t in teacher_field.replace(',', ';').split(';') if t.strip()]
                    elif isinstance(teacher_field, (list, tuple)):
                        teachers = list(teacher_field)
                    else:
                        teachers = [str(teacher_field)]

                    for teacher in teachers:
                        if teacher not in teacher_schedules:
                            teacher_schedules[teacher] = {}
                        if day not in teacher_schedules[teacher]:
                            teacher_schedules[teacher][day] = {}

                        teacher_schedules[teacher][day][lesson_num] = {
                            'subject': lesson_data.get('subject', ''),
                            'teacher': teacher,
                            'group': group_name,
                            'color_bg': lesson_data.get('color_bg', '#FFFFFF'),
                            'color_fg': lesson_data.get('color_fg', '#000000')
                        }
        
        # Save to Excel
        if 'Teacher Schedules' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Teacher Schedules')
            ws.append(['Teacher', 'Day', 'Lesson', 'Subject', 'Group', 'Color BG', 'Color FG'])
        else:
            ws = self.wb['Teacher Schedules']
            ws.delete_rows(2, ws.max_row)
        
        for teacher, schedule in teacher_schedules.items():
            for day, lessons in schedule.items():
                for lesson_num, lesson_data in lessons.items():
                    ws.append([
                        teacher,
                        day,
                        lesson_num,
                        lesson_data.get('subject', ''),
                        lesson_data.get('group', ''),
                        lesson_data.get('color_bg', '#FFFFFF'),
                        lesson_data.get('color_fg', '#000000')
                    ])
        
        self.save()
    
    def clear_all_schedules(self):
        """Clear all schedules (group and teacher)"""
        if 'Group Schedules' in self.wb.sheetnames:
            ws = self.wb['Group Schedules']
            ws.delete_rows(2, ws.max_row)
        
        if 'Teacher Schedules' in self.wb.sheetnames:
            ws = self.wb['Teacher Schedules']
            ws.delete_rows(2, ws.max_row)
        
        self.save()
    
    def save(self):
        """Save Excel file"""
        self.wb.save(self.file_path)
